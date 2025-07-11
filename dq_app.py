import streamlit as st
import pandas as pd
import snowflake.connector
import re
import io
from openpyxl import Workbook
from dotenv import load_dotenv
import os
from openai import AuthenticationError, RateLimitError, OpenAIError

# Load environment variables if present
load_dotenv()

def regex_email_rate(series):
    email_pattern = re.compile(r"^[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,}$")
    if series.empty:
        return 0.0
    def clean_email(val):
        # Remove all unicode whitespace
        return re.sub(r'\s+', '', str(val)).strip()
    cleaned = series.dropna().apply(clean_email)
    print("[DEBUG] Checking email regex on values:")
    for val in cleaned:
        print(f"  Value: {repr(val)} | Match: {bool(email_pattern.match(val))} | ASCII: {[ord(c) for c in val]}")
    # Direct test
    test_val = 'andersoncody@gmail.com'
    print("[DEBUG] Direct regex test on 'andersoncody@gmail.com':", bool(email_pattern.match(test_val)))
    match_count = cleaned.apply(lambda x: bool(email_pattern.match(x))).sum()
    return match_count / len(cleaned) if len(cleaned) > 0 else 0.0

def get_snowflake_connection(account, user, password, warehouse, database, schema):
    ctx = snowflake.connector.connect(
        account=account,
        user=user,
        password=password,
        warehouse=warehouse,
        database=database,
        schema=schema
    )
    return ctx

def get_tables(conn, database, schema):
    query = f"""
        SELECT TABLE_NAME
        FROM {database}.INFORMATION_SCHEMA.TABLES
        WHERE TABLE_SCHEMA = '{schema}'
    """
    cur = conn.cursor()
    cur.execute(query)
    tables = [row[0] for row in cur.fetchall()]
    cur.close()
    return tables

def get_table_metrics(conn, database, schema, table, chunk_size=10000):
    # Get column names (for metrics, not for data fetch)
    cur = conn.cursor()
    cur.execute(f"SELECT COLUMN_NAME, DATA_TYPE FROM {database}.INFORMATION_SCHEMA.COLUMNS WHERE TABLE_SCHEMA = '{schema}' AND TABLE_NAME = '{table}'")
    columns_for_metrics = [row[0] for row in cur.fetchall()]
    cur.close()
    
    # Get row count
    cur = conn.cursor()
    cur.execute(f"SELECT COUNT(*) FROM {database}.{schema}.{table}")
    row_count = cur.fetchone()[0]
    cur.close()
    
    # If table is empty, return empty metrics
    if row_count == 0:
        return [{
            'table': table,
            'column': col,
            'row_count': 0,
            'null_%': None,
            'distinct_count': None,
            'min': None,
            'max': None,
            'email_regex_rate': None
        } for col in columns_for_metrics]
    
    # Fetch all data (or in chunks if large)
    data = []
    offset = 0
    columns = None
    while offset < row_count:
        cur = conn.cursor()
        cur.execute(f"SELECT * FROM {database}.{schema}.{table} LIMIT {chunk_size} OFFSET {offset}")
        chunk = cur.fetchall()
        if columns is None:
            columns = [desc[0] for desc in cur.description]  # Get correct column order
        if not chunk:
            break
        data.extend(chunk)
        cur.close()
        offset += chunk_size
        if offset >= 100000:  # Limit to 100k rows for performance
            break
    df = pd.DataFrame(data, columns=columns)
    
    metrics = []
    # Print all column names for debugging (with repr)
    print("[DEBUG] Columns in DataFrame (repr):", [repr(col) for col in columns])
    for col in columns:
        print(f"[DEBUG] Sample values for column {repr(col)}:", df[col].dropna().head(5).tolist())
        col_data = df[col]
        null_pct = col_data.isnull().mean() * 100
        distinct_count = col_data.nunique(dropna=True)
        min_val = col_data.min() if pd.api.types.is_numeric_dtype(col_data) or pd.api.types.is_datetime64_any_dtype(col_data) else None
        max_val = col_data.max() if pd.api.types.is_numeric_dtype(col_data) or pd.api.types.is_datetime64_any_dtype(col_data) else None
        email_rate = None
        # Only apply regex if the column is 'email' (case-insensitive, strip spaces)
        if col.strip().lower() == "email":
            print(f"[DEBUG] Sample values for column '{col}':", col_data.dropna().tolist())
            email_rate = regex_email_rate(col_data)
        metrics.append({
            'table': table,
            'column': col,
            'row_count': row_count,
            'null_%': round(null_pct, 2),
            'distinct_count': distinct_count,
            'min': min_val,
            'max': max_val,
            'email_regex_rate': round(email_rate, 2) if email_rate is not None else None
        })
    return metrics

def main():
    st.title("Snowflake Data Quality Assessment Tool")
    st.write("Scan a Snowflake schema for data quality metrics or upload a CSV for mock/test data.")

    uploaded_file = st.file_uploader("Upload a CSV file for mock/test data", type=["csv"])
    use_mock = uploaded_file is not None

    if use_mock:
        df = pd.read_csv(uploaded_file)
        st.write("Preview of uploaded data:", df.head())
        # Run data quality checks on uploaded DataFrame
        all_metrics = []
        table = "Uploaded CSV"
        row_count = len(df)
        columns = df.columns.tolist()
        for col in columns:
            col_data = df[col]
            null_pct = col_data.isnull().mean() * 100
            distinct_count = col_data.nunique(dropna=True)
            min_val = col_data.min() if pd.api.types.is_numeric_dtype(col_data) or pd.api.types.is_datetime64_any_dtype(col_data) else None
            max_val = col_data.max() if pd.api.types.is_numeric_dtype(col_data) or pd.api.types.is_datetime64_any_dtype(col_data) else None
            email_rate = None
            if col.strip().lower() == "email":
                st.write(f"[DEBUG] Sample values for column '{col}':", col_data.dropna().tolist())
                email_rate = regex_email_rate(col_data)
            all_metrics.append({
                'table': table,
                'column': col,
                'row_count': row_count,
                'null_%': round(null_pct, 2),
                'distinct_count': distinct_count,
                'min': min_val,
                'max': max_val,
                'email_regex_rate': round(email_rate, 2) if email_rate is not None else None
            })
        df_metrics = pd.DataFrame(all_metrics)
        # Fix ArrowInvalid: ensure 'min' and 'max' columns are string type
        for col in ['min', 'max']:
            if col in df_metrics.columns:
                df_metrics[col] = df_metrics[col].astype(str)
        # Streamlit conditional formatting for null_% > 10 and email_regex_rate < 95
        def highlight_null(val):
            try:
                return 'background-color: yellow' if float(val) > 10 else ''
            except:
                return ''
        def highlight_email(val):
            try:
                return 'background-color: yellow' if val is not None and float(val) < 0.95 else ''
            except:
                return ''
        styled_df = df_metrics.style.applymap(highlight_null, subset=['null_%'])
        if 'email_regex_rate' in df_metrics.columns:
            styled_df = styled_df.applymap(highlight_email, subset=['email_regex_rate'])
        st.write(styled_df)
        # Download buttons
        csv = df_metrics.to_csv(index=False).encode('utf-8')
        st.download_button("Download as CSV", data=csv, file_name="dq_metrics.csv", mime="text/csv")
        excel_buffer = io.BytesIO()
        with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
            df_metrics.to_excel(writer, index=False, sheet_name='DQ Metrics')
            worksheet = writer.sheets['DQ Metrics']
            from openpyxl.styles import PatternFill
            yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
            # Null % formatting
            null_col_idx = list(df_metrics.columns).index('null_%') + 1  # 1-based index
            for row in range(2, len(df_metrics) + 2):  # skip header
                cell = worksheet.cell(row=row, column=null_col_idx)
                try:
                    if cell.value is not None and float(cell.value) > 10:
                        cell.fill = yellow_fill
                except:
                    pass
            # Email regex rate formatting
            if 'email_regex_rate' in df_metrics.columns:
                email_col_idx = list(df_metrics.columns).index('email_regex_rate') + 1
                for row in range(2, len(df_metrics) + 2):
                    cell = worksheet.cell(row=row, column=email_col_idx)
                    try:
                        if cell.value is not None and float(cell.value) < 0.95:
                            cell.fill = yellow_fill
                    except:
                        pass
        st.download_button("Download as Excel", data=excel_buffer.getvalue(), file_name="dq_metrics.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        # AI LLM summary
        enable_ai = st.checkbox("Enable AI Summary (OpenAI)", value=True)
        openai_api_key = st.text_input("OpenAI API Key", type="password", value=os.getenv("OPENAI_API_KEY", ""), help="Your OpenAI key is used only for this session.") if enable_ai else None
        if enable_ai and openai_api_key:
            st.markdown("---")
            st.info("Generating AI summary with OpenAI...")
            import llm_advisor
            sample_rows = df_metrics.sample(min(10, len(df_metrics))).to_dict(orient="records")
            ai_summary = llm_advisor.summarize_data_quality(df_metrics, sample_rows, api_key=openai_api_key, save_to_file=True)
            st.markdown("### AI Data Quality Summary")
            st.markdown(ai_summary)
            st.download_button("Download AI Summary (Markdown)", data=ai_summary, file_name="ai_summary.md", mime="text/markdown")
        elif enable_ai:
            st.warning("OpenAI API key not provided. AI summary skipped.")
        return

    with st.form("snowflake_form"):
        account = st.text_input("Snowflake Account URL", value=os.getenv("SNOWFLAKE_ACCOUNT", ""))
        user = st.text_input("Username", value=os.getenv("SNOWFLAKE_USER", ""))
        password = st.text_input("Password or Token", type="password", value=os.getenv("SNOWFLAKE_PASSWORD", ""))
        warehouse = st.text_input("Warehouse", value=os.getenv("SNOWFLAKE_WAREHOUSE", ""))
        database = st.text_input("Database", value=os.getenv("SNOWFLAKE_DATABASE", ""))
        schema = st.text_input("Schema", value=os.getenv("SNOWFLAKE_SCHEMA", ""))
        st.markdown("---")
        enable_ai = st.checkbox("Enable AI Summary (OpenAI)", value=True)
        openai_api_key = st.text_input("OpenAI API Key", type="password", value=os.getenv("OPENAI_API_KEY", ""), help="Your OpenAI key is used only for this session.") if enable_ai else None
        submitted = st.form_submit_button("Scan")

    if submitted:
        missing_fields = []
        if not account:
            missing_fields.append("Snowflake Account URL")
        if not user:
            missing_fields.append("Username")
        if not password:
            missing_fields.append("Password or Token")
        if not warehouse:
            missing_fields.append("Warehouse")
        if not database:
            missing_fields.append("Database")
        if not schema:
            missing_fields.append("Schema")
        if missing_fields:
            st.error("Please fill in the following fields: " + ", ".join(missing_fields))
            return
        try:
            with st.spinner("Connecting to Snowflake..."):
                conn = get_snowflake_connection(account, user, password, warehouse, database, schema)
            st.success("Connected!")
            tables = get_tables(conn, database, schema)
            if not tables:
                st.warning("No tables found in this schema.")
                return
            all_metrics = []
            for table in tables:
                st.info(f"Scanning table: {table}")
                metrics = get_table_metrics(conn, database, schema, table)
                all_metrics.extend(metrics)
            df_metrics = pd.DataFrame(all_metrics)

            # Fix ArrowInvalid: ensure 'min' and 'max' columns are string type
            for col in ['min', 'max']:
                if col in df_metrics.columns:
                    df_metrics[col] = df_metrics[col].astype(str)

            # Streamlit conditional formatting for null_% > 10 and email_regex_rate < 95
            def highlight_null(val):
                try:
                    return 'background-color: yellow' if float(val) > 10 else ''
                except:
                    return ''
            def highlight_email(val):
                try:
                    return 'background-color: yellow' if val is not None and float(val) < 0.95 else ''
                except:
                    return ''
            styled_df = df_metrics.style.applymap(highlight_null, subset=['null_%'])
            if 'email_regex_rate' in df_metrics.columns:
                styled_df = styled_df.applymap(highlight_email, subset=['email_regex_rate'])
            st.write(styled_df)

            # Download buttons
            csv = df_metrics.to_csv(index=False).encode('utf-8')
            st.download_button("Download as CSV", data=csv, file_name="dq_metrics.csv", mime="text/csv")

            # Excel with conditional formatting
            excel_buffer = io.BytesIO()
            with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
                df_metrics.to_excel(writer, index=False, sheet_name='DQ Metrics')
                worksheet = writer.sheets['DQ Metrics']
                from openpyxl.styles import PatternFill
                yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
                # Null % formatting
                null_col_idx = list(df_metrics.columns).index('null_%') + 1  # 1-based index
                for row in range(2, len(df_metrics) + 2):  # skip header
                    cell = worksheet.cell(row=row, column=null_col_idx)
                    try:
                        if cell.value is not None and float(cell.value) > 10:
                            cell.fill = yellow_fill
                    except:
                        pass
                # Email regex rate formatting
                if 'email_regex_rate' in df_metrics.columns:
                    email_col_idx = list(df_metrics.columns).index('email_regex_rate') + 1
                    for row in range(2, len(df_metrics) + 2):
                        cell = worksheet.cell(row=row, column=email_col_idx)
                        try:
                            if cell.value is not None and float(cell.value) < 0.95:
                                cell.fill = yellow_fill
                        except:
                            pass
            st.download_button("Download as Excel", data=excel_buffer.getvalue(), file_name="dq_metrics.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

            # AI LLM summary
            if enable_ai and openai_api_key:
                st.markdown("---")
                st.info("Generating AI summary with OpenAI...")
                import llm_advisor
                sample_rows = df_metrics.sample(min(10, len(df_metrics))).to_dict(orient="records")
                ai_summary = llm_advisor.summarize_data_quality(df_metrics, sample_rows, api_key=openai_api_key, save_to_file=True)
                st.markdown("### AI Data Quality Summary")
                st.markdown(ai_summary)
                st.download_button("Download AI Summary (Markdown)", data=ai_summary, file_name="ai_summary.md", mime="text/markdown")
            elif enable_ai:
                st.warning("OpenAI API key not provided. AI summary skipped.")
        except Exception as e:
            st.error(f"Error: {e}")

if __name__ == "__main__":
    main() 